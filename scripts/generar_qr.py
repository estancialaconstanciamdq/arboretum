"""
generar_qr.py
Genera un código QR por árbol, apuntando a la URL pública de su ficha.
Los slugs coinciden con los de scripts/cargar_datos.R (misma lógica).

Uso:  python scripts/generar_qr.py
Salida: qr/<slug>.png

Volvé a ejecutarlo si agregás árboles o cambiás nombres/fichas.
"""
import os, re, unicodedata
import openpyxl
import qrcode

BASE_URL = "https://estancialaconstanciamdq.github.io/arboretumdatalab/fichas/"
ARCHIVO  = "data/Arboretum_Master.xlsx"
SALIDA   = "qr"

def slugify(x):
    x = unicodedata.normalize("NFKD", str(x)).encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]+", "-", x).strip("-")

def has(v):
    return v is not None and str(v).strip() not in ("", "-", "...", "None")

def main():
    os.makedirs(SALIDA, exist_ok=True)
    wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
    ws = [wb[s] for s in wb.sheetnames if "rbol" in s.lower()][0]
    rows = list(ws.iter_rows(values_only=True))
    h = rows[0]; data = rows[1:]
    ci = {c: i for i, c in enumerate(h)}
    def g(r, c): return r[ci[c]]

    base = []
    for r in data:
        if has(g(r, "ficha")):        b = slugify(g(r, "ficha"))
        elif has(g(r, "nombre_comun")): b = slugify(g(r, "nombre_comun"))
        else:                          b = str(g(r, "ID")).lower()
        base.append(b)
    from collections import Counter
    cnt = Counter(base)

    n = 0
    for i, r in enumerate(data):
        b = base[i]
        slug = b if cnt[b] == 1 else b + "-" + str(g(r, "ID")).lower()
        img = qrcode.make(BASE_URL + slug + ".html")
        img.save(os.path.join(SALIDA, slug + ".png"))
        n += 1
    print(f"generar_qr.py: {n} códigos QR generados en '{SALIDA}/'.")

if __name__ == "__main__":
    main()
