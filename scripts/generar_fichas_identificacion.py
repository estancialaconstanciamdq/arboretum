"""
generar_fichas_identificacion.py
Genera la ficha de observación (aves y árboles) en LaTeX, a partir de la base
maestra. Reemplaza la edición manual: la lista de especies, los nombres y las
fotos salen de la base. Las imágenes se dimensionan a una ALTURA UNIFORME, así
no hay que ajustar 'scale' foto por foto.

Uso:   python scripts/generar_fichas_identificacion.py
Salida: fichas_identificacion.tex   (compilar con pdflatex)

Se incluyen las especies que tienen foto cargada en la base.
"""
import os, re
import openpyxl

ARCHIVO = "data/Arboretum_Master.xlsx"
SALIDA_AVES = "identificacion_aves.tex"
SALIDA_ARB  = "identificacion_arboles.tex"
DIR_FOTOS = "fotos"          # fotos del sitio (columna 'foto')
DIR_IDENT = "identificacion" # fotos propias de la ficha (columna 'foto_ident')
LOGO = "fotos/logo_lc.jpg"   # logo del encabezado (si existe)
POR_FILA  = 5
FILAS_PAG = 2                # 2 filas de 5 = 10 por página
ALTURA_IMG = "3cm"           # altura uniforme de cada foto

def has(v): return v is not None and str(v).strip() not in ("", "-", "...", "None")

def esc(s):
    s = str(s)
    for a, b in [("\\", r"\textbackslash{}"), ("&", r"\&"), ("%", r"\%"),
                 ("#", r"\#"), ("_", r"\_"), ("$", r"\$"), ("{", r"\{"), ("}", r"\}"),
                 ("×", r"$\times$")]:
        s = s.replace(a, b)
    return s.strip()

def hoja(wb, patron):
    for s in wb.sheetnames:
        if re.search(patron, s, re.I):
            return wb[s]
    raise SystemExit("No se encontró la hoja: " + patron)

def leer(ws, col_nombre, col_sci, col_foto):
    rows = list(ws.iter_rows(values_only=True))
    h = rows[0]; ci = {c: i for i, c in enumerate(h)}
    def g(r, c): return r[ci[c]] if c in ci else None
    out = []
    for r in rows[1:]:
        foto  = g(r, col_foto)
        ident = g(r, "foto_ident")   # imagen propia de la ficha (opcional)
        nom   = g(r, col_nombre)
        sci   = g(r, col_sci)
        # se incluye si hay alguna imagen y algún nombre
        if (has(foto) or has(ident)) and (has(nom) or has(sci)):
            out.append({
                "nombre": esc(nom) if has(nom) else "",
                "sci": esc(sci) if has(sci) else "",
                "foto": str(foto).strip() if has(foto) else "",
                "ident": str(ident).strip() if has(ident) else ""
            })
    return out

def celda(item):
    # preferir la imagen propia de la ficha; si no, usar la foto del sitio
    ruta = f"{DIR_IDENT}/{item['ident']}" if item["ident"] else f"{DIR_FOTOS}/{item['foto']}"
    img = f"\\includegraphics[height={ALTURA_IMG},keepaspectratio]{{{ruta}}}"
    linea2 = item["nombre"]
    if item["sci"]:
        linea2 += (" \\\\ " if item["nombre"] else "") + "\\textit{" + item["sci"] + "}"
    return ("\\begin{minipage}{0.19\\textwidth}\n\\centering\n"
            f"{img}\\\\[0.3em]\n{linea2} \\quad \\LARGE$\\square$\n\\end{{minipage}}")

def bloque(titulo, items, agradecimiento):
    partes = []
    n = POR_FILA * FILAS_PAG
    paginas = [items[i:i+n] for i in range(0, len(items), n)] or [[]]
    logo = (f"\\includegraphics[width=4cm]{{{LOGO}}}\\\\[0.4em]\n"
            if os.path.exists(LOGO) else "")
    for pi, pag in enumerate(paginas):
        partes.append("\\begin{center}\n" + logo + "{\\LARGE \\textbf{" + titulo + "}}\n\\end{center}")
        partes.append("\\vspace{0.2em}\n{\\large Marcá con un tick ($\\checkmark$) las especies que logres identificar.}\n\\vspace{0.8em}")
        filas = [pag[i:i+POR_FILA] for i in range(0, len(pag), POR_FILA)]
        for fi, fila in enumerate(filas):
            partes.append("\\begin{center}\n" + "\n\\hfill\n".join(celda(x) for x in fila) + "\n\\end{center}")
            if fi < len(filas) - 1:
                partes.append("\\vspace{1.8em}")
        partes.append("\\vfill\n\\begin{center}\\scriptsize " + agradecimiento + "\\end{center}")
        partes.append("\\newpage")
    return "\n".join(partes)

def main():
    wb = openpyxl.load_workbook(ARCHIVO, data_only=True)
    aves    = leer(hoja(wb, "aves"),  "especie",      "nombre_cientifico", "foto")
    arboles = leer(hoja(wb, "rbol"),  "nombre_comun", "nombre_cientifico", "foto")

    agr_aves = ("Agradecemos a Carola Grassi, Nahuel Pascual, al equipo del relevamiento "
                "ecológico realizado en 2020 y al equipo de guías por las fotografías e "
                "información. Estancia La Constancia, Argentina.")
    agr_arb  = ("Agradecemos al equipo de guías por las fotografías e información utilizadas "
                "en este material didáctico. Estancia La Constancia, Argentina.")

    PREAMBULO = r"""\documentclass[11pt]{article}
\usepackage[spanish]{babel}
\usepackage[utf8]{inputenc}
\usepackage[T1]{fontenc}
\usepackage{lmodern}
\usepackage{graphicx}
\usepackage{amssymb}
\usepackage{geometry}
\geometry{a4paper, landscape, margin=1.2cm}
\pagestyle{empty}
\begin{document}
"""
    def documento(titulo, items, agr):
        cuerpo = bloque(titulo, items, agr)
        doc = PREAMBULO + cuerpo + "\n\\end{document}\n"
        return doc.replace("\\newpage\n\\end{document}", "\\end{document}")

    with open(SALIDA_AVES, "w", encoding="utf-8") as f:
        f.write(documento("FICHA DE OBSERVACIÓN DE AVES", aves, agr_aves))
    with open(SALIDA_ARB, "w", encoding="utf-8") as f:
        f.write(documento("FICHA DE OBSERVACIÓN DE PLANTAS", arboles, agr_arb))

    print(f"generar_fichas_identificacion.py: generados {SALIDA_AVES} "
          f"({len(aves)} aves) y {SALIDA_ARB} ({len(arboles)} árboles).")

if __name__ == "__main__":
    main()
